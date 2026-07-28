#!/usr/bin/env python3
"""
Import a compatibility chart into PostgreSQL (reactive_groups + compatibility).

TWO source layouts are supported and auto-detected:

  LIST   — one row per reactive group with an "Incompatible With (Group Nos)"
           column (e.g. the 46 CFR Part 150 export, the Odfjell chart).

  MATRIX — a coloured 2-D grid where each cell is an X / blank intersection of a
           row group and a column group, and the CELL FILL COLOUR encodes
           exceptions (e.g. "Cargo Library 3 TABLE OF CHEMICAL CARGO.xlsx"):

               yellow  X   -> incompatible                       (compatible=NO)
               olive   X   -> incompatible, but an EXCEPTION      (NO  + note, Annex I(b))
               blue  cell  -> compatible, but an EXCEPTION/note   (YES + note, Annex I(a))
               dark-blue   -> "other exception" (the code-44 column, line 44)

Both layouts are reduced to the same intermediate form:

    groups  : [(group_code, group_name), ...]
    records : [(code_a, code_b, "YES"|"NO", reaction_description|None), ...]

and share one insert path (source -> reactive_groups -> compatibility), all
inside a single transaction (commit on success, rollback on error).

Pipeline:
    1. get_or_create_source()      resolve the source row from the file name
    2. load_chart()                detect layout, parse -> (groups, records, meta)
    3. insert_reactive_groups()    idempotent insert -> {code: db_id} mapping
    4. insert_compatibilities()    idempotent insert (dedupe on a/b/source)

Usage:
    python cargo_compatibility.py
    python cargo_compatibility.py --file "/path/to/chart.xlsx"
    python cargo_compatibility.py --format matrix        # force a layout
    python cargo_compatibility.py --dry-run

Reads DATABASE_URL from the .env file in this directory.
"""

import argparse
import logging
import os
import re
import sys
from pathlib import Path

from _paths import input_file
from typing import Dict, List, Optional, Set, Tuple

import openpyxl
import pandas as pd
import psycopg2
from psycopg2.extras import execute_values
from dotenv import load_dotenv

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
DEFAULT_FILE = input_file("Cargo Library 3 TABLE OF CHEMICAL CARGO.xlsx")
# DEFAULT_FILE = input_file("Cargo Library 4 Odfjell - Compatibility Chart and Notes Reactive Cargoes - 1999 .xlsx")

SYSTEM = "46 CFR Part 150"                     # reactive_groups.system (the group taxonomy)
HEADER_NO_COL = "reactive group no"            # header cell that marks a LIST-layout table
LIST_CHART_SHEET = "Compatibility Chart"       # preferred sheet name for LIST xlsx files

# compatibility.compatible is a Boolean: X in the chart => incompatible (False),
# a blank cell => compatible (True).
COMPATIBLE = True
INCOMPATIBLE = False

# reactive_groups.group_type: codes 1–22 are REACTIVE groups, 30–43 (and the
# "other exceptions" pseudo-group 44) are CARGO groups.
GROUP_TYPE_REACTIVE = "REACTIVE"
GROUP_TYPE_CARGO = "CARGO"


def group_type_for(code: str) -> str:
    """REACTIVE for codes 1–22, CARGO otherwise (30–43, 44)."""
    try:
        return GROUP_TYPE_REACTIVE if 1 <= int(code) <= 22 else GROUP_TYPE_CARGO
    except (TypeError, ValueError):
        return GROUP_TYPE_CARGO

# MATRIX cell fill colours (ARGB) -> meaning.
FILL_YELLOW = "FFFFFF99"        # X: plain incompatible
FILL_OLIVE = "FFCCFFCC"         # X: incompatible, exception (Annex I(b))
FILL_BLUE = "FF00FFFF"          # note: compatible, exception (Annex I(a))
FILL_DARKBLUE = "FF3366FF"      # "other exception" (code-44 column / line 44)

NOTE_INCOMPAT_EXCEPTION = "Exception to incompatibility (46 CFR Part 150 Annex I(b))"
NOTE_COMPAT_EXCEPTION = "Exception to compatibility (46 CFR Part 150 Annex I(a))"
NOTE_OTHER_EXCEPTION = "Other exception (46 CFR Part 150, line 44)"

OTHER_EXCEPTION_CODE = "44"     # pseudo column that flags "other exceptions"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("cargo_compatibility")


# ===========================================================================
# Database & source resolution
# ===========================================================================
def get_db_connection() -> psycopg2.extensions.connection:
    """Open a PostgreSQL connection using DATABASE_URL from the local .env."""
    load_dotenv(Path(__file__).parent.parent / ".env")
    db_url = os.getenv("DATABASE_URL")
    if not db_url:
        sys.exit("Error: DATABASE_URL not set (checked .env).")
    return psycopg2.connect(db_url)


def derive_source_name(path: Path) -> str:
    """Turn a file name into a human-readable source name (split at first ext)."""
    name = path.name
    for ext in (".xlsx", ".xls", ".csv"):
        if ext in name.lower():
            name = name[: name.lower().index(ext)]
            break
    return name.replace("_", " ").strip()


def _norm_name(s: str) -> str:
    """Lowercase, unify dashes, collapse whitespace — for exact fuzzy compare."""
    s = s.lower().replace("—", "-").replace("–", "-")
    return re.sub(r"\s+", " ", s).strip()


def _alnum(s: str) -> str:
    """Reduce to space-separated lowercase alphanumeric tokens — for partial compare."""
    return re.sub(r"[^a-z0-9]+", " ", s.lower()).strip()


def find_source(cur, name: str) -> Optional[Tuple[int, str]]:
    """Return (id, name) of an existing source matching `name`, or None.

    Two passes, so a duplicate is never created when the source is already there:
      1. exact fuzzy    — same name ignoring case/dashes/whitespace
      2. partial match  — one name's alphanumeric text contains the other's
                          (either direction); the LONGEST such match wins so a
                          short generic name can't shadow a more specific one.
    """
    cur.execute("SELECT id, name FROM source")
    rows = cur.fetchall()

    target_norm = _norm_name(name)
    for sid, sname in rows:
        if _norm_name(sname) == target_norm:
            return sid, sname

    target_alnum = _alnum(name)
    best: Optional[Tuple[int, int, str]] = None   # (match_len, id, name)
    for sid, sname in rows:
        ns = _alnum(sname)
        if ns and (ns in target_alnum or target_alnum in ns):
            if best is None or len(ns) > best[0]:
                best = (len(ns), sid, sname)
    if best is not None:
        return best[1], best[2]
    return None


def get_or_create_source(cur, path: Path) -> int:
    """Reuse an existing source (exact or partial name match), else create it."""
    name = derive_source_name(path)

    match = find_source(cur, name)
    if match:
        log.info("Source found: id=%s (%r) [matched %r]", match[0], match[1], name)
        return match[0]

    cur.execute(
        "INSERT INTO source (name, source_type, file_path, notes, "
        "date_ingested, created_at, updated_at) "
        "VALUES (%s, %s, %s, %s, now(), now(), now()) RETURNING id",
        (name, "regulatory", str(path), "Auto-created by cargo_compatibility.py"),
    )
    source_id = cur.fetchone()[0]
    log.info("Source created: id=%s (%r)", source_id, name)
    return source_id


# ===========================================================================
# Small helpers
# ===========================================================================
def _text(value) -> str:
    """Trim a cell to a string ('' for None/NaN/blank)."""
    if value is None:
        return ""
    s = str(value).strip()
    return "" if s.lower() in ("nan", "nat") else s


def _as_code(value) -> Optional[str]:
    """Interpret a cell as a reactive-group code (int 1..44) -> '12', else None."""
    s = _text(value)
    if not s:
        return None
    try:
        n = int(float(s))
    except (TypeError, ValueError):
        return None
    return str(n) if 1 <= n <= 44 else None


# ===========================================================================
# LIST layout  (Reactive Group No | Name | Incompatible With (Group Nos))
# ===========================================================================
def read_list(path: Path) -> Tuple[List[Tuple[str, str]], List[Tuple[str, str, bool, Optional[str]]], dict]:
    """Parse the LIST layout into (groups, records, meta).

    Each group's 'Incompatible With' cell lists the codes it clashes with; the
    full symmetric matrix is generated: (a, b) is NO if either lists the other.
    """
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xls"):
        try:
            raw = pd.read_excel(path, sheet_name=LIST_CHART_SHEET, header=None, dtype=str)
        except Exception:
            raw = pd.read_excel(path, sheet_name=0, header=None, dtype=str)
    elif suffix == ".csv":
        raw = pd.read_csv(path, header=None, dtype=str, keep_default_na=False)
    else:
        raise ValueError(f"Unsupported file type {suffix!r}; use .csv or .xlsx.")

    header_idx = next(
        (i for i in range(len(raw)) if _text(raw.iat[i, 0]).lower() == HEADER_NO_COL),
        None,
    )
    if header_idx is None:
        raise ValueError(f"Could not find a {HEADER_NO_COL!r} header row in {path.name}.")

    groups: List[Tuple[str, str]] = []
    seen: Set[str] = set()
    incompatible: Dict[str, Set[str]] = {}
    for i in range(header_idx + 1, len(raw)):
        code = _text(raw.iat[i, 0])
        if not code.isdigit() or code in seen:
            continue
        seen.add(code)
        name = _text(raw.iat[i, 1]) if raw.shape[1] > 1 else ""
        incompat_raw = _text(raw.iat[i, 2]) if raw.shape[1] > 2 else ""
        groups.append((code, name))
        incompatible[code] = {t.strip() for t in incompat_raw.split(",") if t.strip().isdigit()}

    codes = [c for c, _ in groups]
    records: List[Tuple[str, str, bool, Optional[str]]] = []
    for a in codes:
        for b in codes:
            if a == b:
                continue
            incompat = (b in incompatible.get(a, set())) or (a in incompatible.get(b, set()))
            records.append((a, b, INCOMPATIBLE if incompat else COMPATIBLE, None))

    log.info("LIST layout: %d groups, %d compatibility cells", len(groups), len(records))
    return groups, records, {}


# ===========================================================================
# MATRIX layout  (coloured 2-D grid)
# ===========================================================================
def _cell_fill(cell) -> Optional[str]:
    """Return a cell's solid fill colour as an upper-case ARGB string, or None."""
    fill = cell.fill
    if fill is None or fill.patternType is None:
        return None
    fg = fill.fgColor
    if fg is None or fg.type != "rgb" or not isinstance(fg.rgb, str):
        return None
    rgb = fg.rgb.upper()
    return None if rgb in ("FFFFFFFF", "00000000") else rgb   # treat white/none as blank


def _classify_cell(value, fill: Optional[str]) -> Tuple[bool, Optional[str]]:
    """Map an (X/blank value, fill colour) to (compatible: bool, reaction_description)."""
    is_x = _text(value).upper() == "X"
    if is_x:
        if fill == FILL_OLIVE:
            return INCOMPATIBLE, NOTE_INCOMPAT_EXCEPTION
        return INCOMPATIBLE, None                         # yellow / uncoloured X
    if fill == FILL_BLUE:
        return COMPATIBLE, NOTE_COMPAT_EXCEPTION
    if fill == FILL_DARKBLUE:
        return COMPATIBLE, NOTE_OTHER_EXCEPTION
    return COMPATIBLE, None


def _grid(ws) -> List[list]:
    """Materialize a worksheet's values as a list of rows (fast, style-free)."""
    return [list(row) for row in ws.iter_rows(values_only=True)]


def _gcell(grid: List[list], r: int, c: int):
    """1-based cell access into a values grid (None when out of range)."""
    if 1 <= r <= len(grid):
        row = grid[r - 1]
        if 1 <= c <= len(row):
            return row[c - 1]
    return None


def _count_x(grid: List[list]) -> int:
    """Count 'X' cells — the hallmark of a compatibility matrix sheet."""
    return sum(1 for row in grid for v in row if _text(v).upper() == "X")


def _col_codes(grid: List[list], c: int) -> Dict[int, str]:
    """All reactive-group codes down column `c`, keyed by row (full set, gaps ok)."""
    return {r: code for r in range(1, len(grid) + 1)
            if (code := _as_code(_gcell(grid, r, c))) is not None}


def _row_codes(grid: List[list], r: int, ncols: int) -> Dict[int, str]:
    """All reactive-group codes along row `r`, keyed by column (full set, gaps ok)."""
    return {c: code for c in range(1, ncols + 1)
            if (code := _as_code(_gcell(grid, r, c))) is not None}


def _detect_matrix(grid: List[list]) -> Optional[dict]:
    """Locate the two code axes of a compatibility matrix, orientation-independent.

    The chart has a VERTICAL code axis (a column of codes = the row groups) and a
    HORIZONTAL code axis (a row of codes = the column groups). Either axis may hold
    the fuller set (e.g. Cargo Library 3 puts 30–44 on the columns; the Odfjell
    chart puts 30–43 on the rows), and both may have gaps (…22, 30…), so we read
    EVERY code on each axis — not just a 1,2,3 run.

    Returns {code_col, row_codes, header_row, col_codes} or None.
    """
    nrows = len(grid)
    ncols = max((len(row) for row in grid), default=0)

    # Vertical axis = column with the most codes (leftmost wins ties).
    vbest = None  # (count, col, {row: code})
    for c in range(1, ncols + 1):
        codes = _col_codes(grid, c)
        if len(codes) >= 3 and (vbest is None or len(codes) > vbest[0]):
            vbest = (len(codes), c, codes)
    if vbest is None:
        return None
    code_col, row_codes = vbest[1], vbest[2]

    # Horizontal axis = the row of codes ABOVE the data (topmost wins ties). Using
    # "above the first data row" excludes any legend/mirror code row below it.
    top_data_row = min(row_codes)
    hbest = None  # (count, row, {col: code})
    for r in range(1, top_data_row):
        codes = _row_codes(grid, r, ncols)
        if len(codes) >= 3 and (hbest is None or len(codes) > hbest[0]):
            hbest = (len(codes), r, codes)
    if hbest is None:
        return None
    header_row, col_codes = hbest[1], hbest[2]
    return {"code_col": code_col, "row_codes": row_codes,
            "header_row": header_row, "col_codes": col_codes}


def _pick_matrix_sheet(wb):
    """Choose the matrix sheet: a detectable code-axis pair AND the most 'X' cells.

    'X' count disambiguates the real chart from a data sheet that merely contains a
    1,2,3,… column; the leftmost/topmost axes make a dedicated matrix sheet win over
    an embedded copy inside a data sheet.
    """
    best = None  # (key, ws, grid, axes)
    for ws in wb.worksheets:
        grid = _grid(ws)
        axes = _detect_matrix(grid)
        if axes is None:
            continue
        x_count = _count_x(grid)
        if x_count == 0:
            continue
        key = (x_count, len(axes["row_codes"]) + len(axes["col_codes"]),
               -axes["code_col"], -axes["header_row"])
        if best is None or key > best[0]:
            best = (key, ws, grid, axes)
    if best is None:
        raise ValueError("No compatibility-matrix sheet (code axes + 'X' cells) found.")
    return best[1], best[2], best[3]


def read_matrix(path: Path):
    """Parse the coloured MATRIX layout into (groups, records, meta)."""
    if path.suffix.lower() not in (".xlsx", ".xls"):
        raise ValueError("MATRIX layout needs an .xlsx (cell colours carry meaning).")

    wb = openpyxl.load_workbook(path)          # styles needed for fill colours
    ws, grid, axes = _pick_matrix_sheet(wb)
    code_col, row_codes = axes["code_col"], axes["row_codes"]
    header_row, col_codes = axes["header_row"], axes["col_codes"]

    # Column-group names live on a text banner row above the code axis; pick the
    # row (above header_row) with the most text cells under the code columns.
    name_row, name_hits = None, 0
    for r in range(1, header_row):
        hits = sum(1 for c in col_codes
                   if _text(_gcell(grid, r, c)) and _as_code(_gcell(grid, r, c)) is None)
        if hits > name_hits:
            name_row, name_hits = r, hits

    # Build the code -> name map from both axes (row-header names preferred).
    names: Dict[str, str] = {}
    if name_row:
        for c, code in col_codes.items():
            nm = _text(_gcell(grid, name_row, c))
            if nm:
                names[code] = nm
    for r, code in row_codes.items():
        nm = _text(_gcell(grid, r, code_col + 1))
        if nm:
            names[code] = nm
    # The "other exceptions" column/row (code 44) carries no name of its own.
    names.setdefault(OTHER_EXCEPTION_CODE, "Other exceptions (line 44)")

    all_codes = set(row_codes.values()) | set(col_codes.values())
    groups = [(code, names.get(code, f"Group {code}"))
              for code in sorted(all_codes, key=lambda x: int(x))]

    # One record per (row group, column group) cell. Code 44 is a real CARGO group:
    # its dark-blue cells become compatibility rows with the "other exception" note.
    records: List[Tuple[str, str, bool, Optional[str]]] = []
    seen_pairs: Set[Tuple[str, str]] = set()        # a code may repeat across columns
    for r, rc in row_codes.items():
        for c, cc in col_codes.items():
            if cc == rc or (rc, cc) in seen_pairs:
                continue                                    # diagonal / duplicate
            seen_pairs.add((rc, cc))
            cell = ws.cell(r, c)
            compatible, note = _classify_cell(cell.value, _cell_fill(cell))
            records.append((rc, cc, compatible, note))

    log.info("MATRIX layout: sheet=%r, %d groups, %d cells (%d with notes)",
             ws.title, len(groups), len(records),
             sum(1 for *_, n in records if n))
    return groups, records, {}


# ===========================================================================
# Layout detection + dispatch
# ===========================================================================
def detect_layout(path: Path) -> str:
    """Return 'list' or 'matrix' by inspecting the file (xlsx w/o a list header = matrix)."""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return "list"
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True):
                if any(_text(v).lower() == HEADER_NO_COL for v in row):
                    return "list"
    except Exception:
        pass
    return "matrix"


def load_chart(path: Path, fmt: str):
    """Parse `path` under the chosen (or detected) layout -> (groups, records, meta)."""
    if fmt == "auto":
        fmt = detect_layout(path)
    log.info("Layout: %s", fmt)
    return read_matrix(path) if fmt == "matrix" else read_list(path)


# ===========================================================================
# Inserts
# ===========================================================================
def insert_reactive_groups(
    cur, groups: List[Tuple[str, str]], source_id: int, dry_run: bool
) -> Dict[str, int]:
    """Insert reactive groups (idempotently) and return {group_code: db_id}."""
    cur.execute(
        "SELECT group_code, id FROM reactive_groups WHERE system=%s AND source_id=%s",
        (SYSTEM, source_id),
    )
    mapping: Dict[str, int] = {code: gid for code, gid in cur.fetchall()}

    inserted = 0
    for code, name in groups:
        if code in mapping:
            continue                                        # already exists -> skip
        if dry_run:
            mapping[code] = -len(mapping) - 1               # fake id for dry-run linking
            inserted += 1
            continue
        cur.execute(
            "INSERT INTO reactive_groups "
            "(system, group_code, group_name, group_type, description, source_id, "
            "created_at, updated_at) "
            "VALUES (%s, %s, %s, %s, NULL, %s, now(), now()) RETURNING id",
            (SYSTEM, code, name, group_type_for(code), source_id),
        )
        mapping[code] = cur.fetchone()[0]
        inserted += 1

    log.info("Inserted %d reactive groups (%d already existed).",
             inserted, len(mapping) - inserted)
    return mapping


def insert_compatibilities(
    cur,
    records: List[Tuple[str, str, bool, Optional[str]]],
    mapping: Dict[str, int],
    source_id: int,
    dry_run: bool,
) -> int:
    """Insert one compatibility row per unordered group pair, in canonical order.

    Compatibility is symmetric, so each pair is stored once with
    group_a_id <= group_b_id. When the two directions disagree, INCOMPATIBLE
    (False) wins; a reaction note from either direction is kept.
    """
    # Collapse to canonical unordered pairs. keyed by (min_id, max_id).
    pairs: Dict[Tuple[int, int], Tuple[bool, Optional[str]]] = {}
    for code_a, code_b, compatible, note in records:
        a_id, b_id = mapping.get(code_a), mapping.get(code_b)
        if a_id is None or b_id is None or a_id == b_id:    # no group / self-pair -> skip
            continue
        key = (a_id, b_id) if a_id <= b_id else (b_id, a_id)
        if key in pairs:
            prev_compat, prev_note = pairs[key]
            compatible = prev_compat and compatible          # False (incompatible) wins
            note = prev_note or note
        pairs[key] = (compatible, note)

    # Skip pairs already present for this source (idempotent re-runs).
    existing: Set[Tuple[int, int]] = set()
    if not dry_run:
        cur.execute(
            "SELECT group_a_id, group_b_id FROM compatibility WHERE source_id=%s",
            (source_id,),
        )
        existing = {(a, b) for a, b in cur.fetchall()}

    to_insert = [
        (a_id, b_id, compatible, None, note, source_id)
        for (a_id, b_id), (compatible, note) in pairs.items()
        if (a_id, b_id) not in existing
    ]

    if dry_run:
        log.info("Would insert %d compatibility records (dry-run).", len(to_insert))
        return len(to_insert)

    if to_insert:
        execute_values(
            cur,
            "INSERT INTO compatibility "
            "(group_a_id, group_b_id, compatible, severity, reaction_description, "
            "source_id, created_at, updated_at) VALUES %s "
            "ON CONFLICT (group_a_id, group_b_id) DO NOTHING",
            to_insert,
            template="(%s, %s, %s, %s, %s, %s, now(), now())",
        )
    log.info("Inserted %d compatibility records.", len(to_insert))
    return len(to_insert)


# ===========================================================================
# Appendix I chemical-pair exceptions  ->  compatibility_exception
# ===========================================================================
def _normalize_syn(text: str) -> str:
    """Match the synonym table's normalized_text: lowercase, strip punctuation, collapse ws."""
    s = text.lower()
    s = re.sub(r"[^\w\s]", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


class CargoResolver:
    """Resolve an Appendix I chemical name to a cargo_chemical id.

    Matching, in order: exact canonical name (case-insensitive), normalized
    canonical name, then any synonym (via synonyms.normalized_text). Returns None
    when nothing matches so the caller can log it for manual review.
    """

    def __init__(self, cur):
        cur.execute("SELECT id, canonical_name FROM cargo_chemical")
        self.by_canon: Dict[str, int] = {}
        self.by_canon_norm: Dict[str, int] = {}
        for cid, nm in cur.fetchall():
            if not nm:
                continue
            self.by_canon.setdefault(nm.strip().lower(), cid)
            self.by_canon_norm.setdefault(_normalize_syn(nm), cid)

        cur.execute(
            "SELECT cs.cargo_id, s.normalized_text "
            "FROM cargo_synonym cs JOIN synonyms s ON s.id = cs.synonym_id"
        )
        self.by_syn: Dict[str, int] = {}
        for cid, norm in cur.fetchall():
            if norm:
                self.by_syn.setdefault(norm, cid)
        log.info("CargoResolver: %d canonical names, %d synonyms loaded",
                 len(self.by_canon), len(self.by_syn))

    def resolve(self, name: str) -> Optional[int]:
        raw = _text(name)
        if not raw:
            return None
        if raw.lower() in self.by_canon:
            return self.by_canon[raw.lower()]
        norm = _normalize_syn(raw)
        return self.by_canon_norm.get(norm) or self.by_syn.get(norm)


def _parse_bool(value, default: Optional[bool] = None) -> Optional[bool]:
    """Interpret a cell as compatible (True) / incompatible (False)."""
    s = _text(value).lower()
    if s in ("no", "false", "f", "0", "x", "incompatible", "not compatible"):
        return False
    if s in ("yes", "true", "t", "1", "compatible"):
        return True
    return default


def read_exception_rows(path: Path) -> List[dict]:
    """Read an Appendix I exceptions file (CSV/XLSX) into normalized dict rows.

    Expected columns (case-insensitive; extras ignored):
        cargo_a, cargo_b, compatible, [exception_type], [appendix], [section],
        [reason], [notes]
    """
    if path.suffix.lower() in (".xlsx", ".xls"):
        df = pd.read_excel(path, dtype=str)
    else:
        df = pd.read_csv(path, dtype=str, keep_default_na=False)
    df.columns = [re.sub(r"\s+", "_", str(c).strip().lower()) for c in df.columns]

    def col(row, *names):
        for n in names:
            if n in row and _text(row[n]):
                return _text(row[n])
        return ""

    rows = []
    for _, r in df.iterrows():
        rows.append({
            "cargo_a": col(r, "cargo_a", "cargo_a_name", "chemical_a", "cargo_1"),
            "cargo_b": col(r, "cargo_b", "cargo_b_name", "chemical_b", "cargo_2"),
            "compatible": col(r, "compatible"),
            "exception_type": col(r, "exception_type", "type"),
            "appendix": col(r, "appendix") or "I",
            "section": col(r, "section"),
            "reason": col(r, "reason"),
            "notes": col(r, "notes"),
        })
    return rows


def import_exceptions(cur, path: Path, source_id: int, dry_run: bool) -> Tuple[int, int]:
    """Import Appendix I chemical-pair exceptions into compatibility_exception.

    Cargo names are resolved case-insensitively (canonical + synonyms). Unresolved
    pairs are logged for manual review, never fatal. Returns (inserted, unmatched).
    """
    rows = read_exception_rows(path)
    log.info("Appendix I: %d exception rows read from %s", len(rows), path.name)
    resolver = CargoResolver(cur)

    existing: Set[Tuple[int, int]] = set()
    if not dry_run:
        cur.execute("SELECT cargo_a_id, cargo_b_id FROM compatibility_exception")
        existing = {(a, b) for a, b in cur.fetchall()}

    to_insert, unmatched = [], 0
    seen: Set[Tuple[int, int]] = set()
    for r in rows:
        a_id = resolver.resolve(r["cargo_a"])
        b_id = resolver.resolve(r["cargo_b"])
        if a_id is None or b_id is None or a_id == b_id:
            missing = [n for n, i in ((r["cargo_a"], a_id), (r["cargo_b"], b_id)) if i is None]
            log.warning("Appendix I: UNMATCHED pair (%r, %r) — needs manual review%s",
                        r["cargo_a"], r["cargo_b"],
                        f" [could not match: {missing}]" if missing else "")
            unmatched += 1
            continue

        compatible = _parse_bool(r["compatible"], default=None)
        if compatible is None:
            log.warning("Appendix I: pair (%r, %r) has no clear 'compatible' value — skipping",
                        r["cargo_a"], r["cargo_b"])
            unmatched += 1
            continue

        # exception_type: explicit if given, else derived from the appendix outcome.
        #   compatible (allowed despite matrix)   -> MORE_COMPATIBLE
        #   incompatible (barred despite matrix)  -> LESS_COMPATIBLE
        etype = r["exception_type"].upper().replace(" ", "_") or (
            "MORE_COMPATIBLE" if compatible else "LESS_COMPATIBLE")
        if etype not in ("MORE_COMPATIBLE", "LESS_COMPATIBLE"):
            etype = "MORE_COMPATIBLE" if compatible else "LESS_COMPATIBLE"

        a, b = (a_id, b_id) if a_id <= b_id else (b_id, a_id)   # canonical order
        if (a, b) in existing or (a, b) in seen:
            continue
        seen.add((a, b))
        to_insert.append((a, b, compatible, etype, r["appendix"] or None,
                          r["section"] or None, r["reason"] or None,
                          r["notes"] or None, source_id))

    if dry_run:
        log.info("Appendix I: would insert %d exceptions (%d unmatched).", len(to_insert), unmatched)
        return len(to_insert), unmatched

    if to_insert:
        execute_values(
            cur,
            "INSERT INTO compatibility_exception "
            "(cargo_a_id, cargo_b_id, compatible, exception_type, appendix, section, "
            "reason, notes, source_id, created_at, updated_at) VALUES %s "
            "ON CONFLICT (cargo_a_id, cargo_b_id) DO NOTHING",
            to_insert,
            template="(%s, %s, %s, %s, %s, %s, %s, %s, %s, now(), now())",
        )
    log.info("Inserted %d compatibility exceptions (%d unmatched, logged).", len(to_insert), unmatched)
    return len(to_insert), unmatched


# ===========================================================================
# Main
# ===========================================================================
def main() -> None:
    parser = argparse.ArgumentParser(
        description="Import a compatibility chart (list or coloured matrix) into PostgreSQL."
    )
    parser.add_argument("file_pos", nargs="?", default=None,
                        help="path to the chart CSV/XLSX (positional alternative to --file)")
    parser.add_argument("--file", default=DEFAULT_FILE, help="path to the chart CSV/XLSX")
    parser.add_argument("--format", choices=("auto", "list", "matrix"), default="auto",
                        help="force a layout (default: auto-detect)")
    parser.add_argument("--appendix", default=None,
                        help="path to an Appendix I chemical-pair exceptions file (CSV/XLSX)")
    parser.add_argument("--dry-run", action="store_true", help="parse + report, write nothing")
    args = parser.parse_args()

    path = Path(args.file_pos or args.file)
    if not path.is_file():
        sys.exit(f"Error: file not found: {path}")
    appendix_path = Path(args.appendix) if args.appendix else None
    if appendix_path and not appendix_path.is_file():
        sys.exit(f"Error: appendix file not found: {appendix_path}")

    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            source_id = get_or_create_source(cur, path)

            groups, records, meta = load_chart(path, args.format)
            mapping = insert_reactive_groups(cur, groups, source_id, args.dry_run)
            n_compat = insert_compatibilities(cur, records, mapping, source_id, args.dry_run)

            n_exc = n_unmatched = 0
            if appendix_path:
                n_exc, n_unmatched = import_exceptions(cur, appendix_path, source_id, args.dry_run)

            if args.dry_run:
                log.info("Dry run: rolling back, nothing written.")
                conn.rollback()
            else:
                conn.commit()
                log.info("Import completed successfully. "
                         "(%d reactive groups, %d compatibility records, "
                         "%d exceptions, %d unmatched)",
                         len(mapping), n_compat, n_exc, n_unmatched)
    except Exception:
        conn.rollback()
        log.exception("Import failed - transaction rolled back.")
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    main()

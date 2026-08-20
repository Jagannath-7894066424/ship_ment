#!/usr/bin/env python3
"""
Build an Excel workbook mapping every data source to the tables that depend on it.

Four sheets:
  Sources        one row per source: identity, ranks, input files, loader scripts,
                 how many tables it populates and how many rows it owns
  Source x Table one row per (source, table) pair - the dependency grid, with the
                 join path and the live row count
  Tables         one row per table: how it is scoped to a source, which sources
                 populate it, total rows
  Input Files    one row per input file: which source it feeds, which loader
                 reads it, which tables that loader writes

Two ways a table depends on a source:
  DIRECT   the table carries a source_id column (18 tables)
  INDIRECT the table has no source_id and inherits it through a parent FK
           (e.g. procedure_template_steps -> procedure_templates.source_id)

Row counts are read live; a table that is locked or slow is reported as None
rather than blocking the whole build.

Usage:
    python3 etl/build_source_table_map.py
    python3 etl/build_source_table_map.py --out path/to/file.xlsx
    python3 etl/build_source_table_map.py --no-counts      # structure only, fast
"""

import argparse
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import psycopg2
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_OUT = REPO_ROOT / "etl" / "data" / "source_table_map.xlsx"

# Tables with no source_id that inherit one through a parent FK.
#   child -> (parent table, child FK column, parent PK column)
INDIRECT: Dict[str, Tuple[str, str, str]] = {
    "procedure_template_steps":       ("procedure_templates", "procedure_templates_id", "id"),
    "procedure_template_instruction": ("procedure_templates", "procedure_templates_id", "id"),
    "cleaning_process_step":          ("cleaning_process", "cleaning_process_id", "id"),
    "cargo_operational_requirement":  ("operational_requirement", "operational_requirement_id", "id"),
    "master_cargo_chemical_group_details": ("reactive_groups", "group_code", "group_code"),
}

# Tables that are deliberately source-agnostic.
UNSCOPED = {
    "source":            "the source registry itself",
    "field_definitions": "shared property vocabulary across all sources",
    "synonyms":          "shared name vocabulary; the link row cargo_synonym carries the source",
    "changelog":         "audit trail, references a source but is not owned by one",
    "coating_company":   "reference data, not source-scoped",
    "coating_system":    "reference data, not source-scoped",
    "marine_chemical_use": "links marine_chemicals to cargo; scope comes from the parent",
}

# source name -> (input files, loader scripts). Mirrors run_all.sh.
SOURCE_PIPELINE: Dict[str, Tuple[List[str], List[str]]] = {
    "Lars Stole Birkeland — Chemical Cargo specifications - 2002": (
        ["Lars Stole Birkeland - Chemical Cargo specifications - 2002.xlsx - CGOSPEC.csv"],
        ["master_loader.py"]),
    "Unknown - Products CHEM - 1996": (
        ["Unknown - Products CHEM - 1996.XLS"], ["master_loader.py"]),
    "USCG Chemical Data Guide For Bulk Shipment By Water [7th Edition 1990]_reviewed": (
        ["USCG Chemical Data Guide For Bulk Shipment By Water [7th Edition 1990]_reviewed.csv"],
        ["master_loader.py"]),
    "USCG CHRIS Chemical Data Guide": (
        ["USCG CHRIS Chemical Data Guides_chemical_exceptions.csv"],
        ["compatibility_exception_loader.py"]),
    "IBC Code": (["IBC Code.xlsx", "Operational_References_Master.csv"],
                 ["master_loader.py", "cargo_operational_requirement.py"]),
    "Miracle Tank Cleaning Guide": (["Miracle Tank Cleaning Guide.xlsx"], ["master_loader.py"]),
    "Sittigs Handbook of Toxic & Hazardous Chemicals": (
        ["Sittigs Handbook of Toxic & Hazardous Chemicals.csv"], ["sittig_handbook.py"]),
    "Dr Verweys Tank Cleaning Guide": (
        ["Dr. Verwey's Tank Cleaning Table 4.xlsx - CLEANING PROCEDURES (T-2).csv",
         "Dr. Verwey's Tank Cleaning Table 4.xlsx - CLEANING chemical_to_chemical.csv"],
        ["proceduretemplate.py", "verwey_cleaning.py"]),
    "Dr Verweys Tank Cleaning Guide Pdf Book": (
        ["Dr Verweys Tank Cleaning Guide Pdf Book - Cargo details.csv",
         "Dr Verweys Tank Cleaning Guide Pdf Book Procdure Template.csv",
         "Dr Verweys Tank Cleaning Guide Pdf Book  _from-to procedure.csv"],
        ["verwey_pdf_book.py", "verwey_pdf_book_procedures.py",
         "verwey_pdf_book_families.py", "verwey_pdf_book_matrix.py"]),
    "Drew Ameroid Tank Cleaning Guide (TCG)": (
        ["Drew Ameroid Tank Cleaning Guide.xlsx"], ["drew_ameroid.py"]),
    "Odfjell Compatibility Chart": (
        ["Cargo Library 4 Odfjell - Compatibility Chart and Notes Reactive Cargoes - 1999 .xlsx",
         "cargo_chemicals_groupData.csv"],
        ["cargo_compatibility.py", "reactive_group.py", "link_cargo_reactive_groups.py"]),
    "Cargo Library 4 Odfjell - Compatibility Chart and Notes Reactive Cargoes - 1999": (
        ["Cargo Library 4 Odfjell - Compatibility Chart and Notes Reactive Cargoes - 1999 .xlsx"],
        ["cargo_compatibility.py"]),
    "Cargo Library 3 TABLE OF CHEMICAL CARGO": (
        ["Cargo Library 3 TABLE OF CHEMICAL CARGO.xlsx"], ["master_loader.py"]),
    "DOT Hazardous Materials Table (49 CFR 172.101)": (
        ["Cargo Library 2 DOT_hazardous_materials.xlsx - HM Table.csv"],
        ["dot_hmt_extract.py", "dot_hazmat_symbol_loader.py", "cargo_dot_hazad_loader.py"]),
}

HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF")
SUB_FILL = PatternFill("solid", fgColor="DDEBF7")


def connect():
    load_dotenv(REPO_ROOT / ".env")
    dsn = os.environ.get("DATABASE_URL")
    if not dsn:
        sys.exit("DATABASE_URL not set (check .env)")
    return psycopg2.connect(dsn)


def safe_count(cur, sql: str, params=()) -> Optional[int]:
    """Row count, or None if the table is locked / the query is too slow."""
    try:
        cur.execute("SET LOCAL statement_timeout='4s'")
        cur.execute(sql, params)
        return cur.fetchone()[0]
    except Exception:
        cur.connection.rollback()
        return None


def gather(cur, with_counts: bool):
    cur.execute("""SELECT table_name FROM information_schema.tables
                   WHERE table_schema='public' AND table_type='BASE TABLE'
                     AND table_name <> '_prisma_migrations' ORDER BY 1""")
    tables = [r[0] for r in cur.fetchall()]

    cur.execute("""SELECT table_name FROM information_schema.columns
                   WHERE table_schema='public' AND column_name='source_id'""")
    direct = {r[0] for r in cur.fetchall()}

    cur.execute("""SELECT id, name, coalesce(edition,''), rank_regulatory, rank_cleaning,
                          rank_compatibility, rank_physical, rank_health
                     FROM source ORDER BY id""")
    sources = cur.fetchall()

    # (source_id, table) -> rows
    grid: Dict[Tuple[int, str], Optional[int]] = {}
    totals: Dict[str, Optional[int]] = {}

    for t in tables:
        totals[t] = safe_count(cur, f'SELECT count(*) FROM "{t}"') if with_counts else None

    if with_counts:
        for t in sorted(direct):
            rows = safe_count(cur, f'SELECT source_id, count(*) FROM "{t}" GROUP BY 1')
            # safe_count only returns one value; re-run properly for the grouped case
            try:
                cur.execute("SET LOCAL statement_timeout='4s'")
                cur.execute(f'SELECT source_id, count(*) FROM "{t}" GROUP BY 1')
                for sid, n in cur.fetchall():
                    if sid is not None:
                        grid[(sid, t)] = n
            except Exception:
                cur.connection.rollback()

        for child, (parent, fk, pk) in INDIRECT.items():
            if child not in tables or parent not in tables:
                continue
            try:
                cur.execute("SET LOCAL statement_timeout='8s'")
                cur.execute(f'''SELECT p.source_id, count(*) FROM "{child}" c
                                JOIN "{parent}" p ON p."{pk}" = c."{fk}"
                                GROUP BY 1''')
                for sid, n in cur.fetchall():
                    if sid is not None:
                        grid[(sid, child)] = n
            except Exception:
                cur.connection.rollback()

    return tables, direct, sources, grid, totals


def style_header(ws, ncols: int) -> None:
    for i in range(1, ncols + 1):
        cell = ws.cell(row=1, column=i)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def autosize(ws, limit: int = 62) -> None:
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 2, limit)


def build(out: Path, with_counts: bool) -> None:
    conn = connect()
    try:
        with conn.cursor() as cur:
            tables, direct, sources, grid, totals = gather(cur, with_counts)
    finally:
        conn.close()

    wb = Workbook()

    # ---- Sheet 1: Sources ------------------------------------------------
    ws = wb.active
    ws.title = "Sources"
    ws.append(["source_id", "source name", "edition", "input file(s)", "loader script(s)",
               "dependent tables", "tables populated", "rows owned",
               "rank_regulatory", "rank_cleaning", "rank_compatibility",
               "rank_physical", "rank_health"])
    for sid, name, edition, r_reg, r_cln, r_cmp, r_phy, r_hlt in sources:
        files, loaders = SOURCE_PIPELINE.get(name, ([], []))
        mine = {t: n for (s, t), n in grid.items() if s == sid and n}
        ws.append([
            sid, name, edition,
            "\n".join(files) or "—",
            "\n".join(loaders) or "—",
            "\n".join(sorted(mine)) or "—",
            len(mine),
            sum(mine.values()) if mine else 0,
            r_reg, r_cln, r_cmp, r_phy, r_hlt,
        ])
    for row in ws.iter_rows(min_row=2):
        for c in (row[3], row[4], row[5]):
            c.alignment = Alignment(wrap_text=True, vertical="top")
    style_header(ws, 13)
    autosize(ws, 46)

    # ---- Sheet 2: Source x Table ----------------------------------------
    ws = wb.create_sheet("Source x Table")
    ws.append(["source_id", "source name", "dependent table", "dependency",
               "join path", "rows for this source"])
    id_to_name = {sid: name for sid, name, *_ in sources}
    for (sid, t), n in sorted(grid.items(), key=lambda kv: (kv[0][0], -(kv[1] or 0))):
        if t in direct:
            kind, path = "DIRECT", f"{t}.source_id"
        else:
            parent, fk, pk = INDIRECT[t]
            kind, path = "INDIRECT", f"{t}.{fk} -> {parent}.{pk} -> {parent}.source_id"
        ws.append([sid, id_to_name.get(sid, "?"), t, kind, path, n])
    style_header(ws, 6)
    autosize(ws)

    # ---- Sheet 3: Tables -------------------------------------------------
    ws = wb.create_sheet("Tables")
    ws.append(["table", "scoping", "join path / reason", "sources populating it",
               "source count", "total rows"])
    for t in tables:
        contributors = sorted(id_to_name.get(s, "?") for (s, tt), n in grid.items()
                              if tt == t and n)
        if t in direct:
            kind, note = "DIRECT", f"{t}.source_id"
        elif t in INDIRECT:
            parent, fk, pk = INDIRECT[t]
            kind, note = "INDIRECT", f"via {parent}.source_id ({t}.{fk})"
        else:
            kind, note = "NOT SCOPED", UNSCOPED.get(t, "no source linkage")
        ws.append([t, kind, note, "\n".join(contributors) or "—",
                   len(contributors), totals.get(t)])
    for row in ws.iter_rows(min_row=2):
        row[3].alignment = Alignment(wrap_text=True, vertical="top")
    style_header(ws, 6)
    autosize(ws, 46)

    # ---- Sheet 4: Input Files -------------------------------------------
    ws = wb.create_sheet("Input Files")
    ws.append(["input file", "source name", "loader script(s)", "tables written"])
    for name, (files, loaders) in sorted(SOURCE_PIPELINE.items()):
        sid = next((s for s, n, *_ in sources if n == name), None)
        written = sorted(t for (s, t), n in grid.items() if s == sid and n)
        for f in files:
            ws.append([f, name, "\n".join(loaders), "\n".join(written) or "—"])
    for row in ws.iter_rows(min_row=2):
        for c in (row[2], row[3]):
            c.alignment = Alignment(wrap_text=True, vertical="top")
    style_header(ws, 4)
    autosize(ws, 52)

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Wrote {out.relative_to(REPO_ROOT)}")
    print(f"  sources           : {len(sources)}")
    print(f"  tables            : {len(tables)}  ({len(direct)} direct, "
          f"{len([t for t in tables if t in INDIRECT])} indirect, "
          f"{len([t for t in tables if t not in direct and t not in INDIRECT])} unscoped)")
    print(f"  source-table pairs: {len(grid)}")
    generated = datetime.now().strftime("%Y-%m-%d %H:%M")
    print(f"  generated         : {generated}")


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--out", default=str(DEFAULT_OUT))
    ap.add_argument("--no-counts", action="store_true",
                    help="skip row counts (fast; use while a load is running)")
    args = ap.parse_args()
    build(Path(args.out), with_counts=not args.no_counts)


if __name__ == "__main__":
    main()
